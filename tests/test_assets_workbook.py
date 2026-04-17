from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook

from src.assets_workbook import parse_assets_workbook


def _append_standard_rows(sheet, display_names: list[str], tickers: list[str]) -> None:
    name_row: list[object] = []
    security_row: list[object] = []
    start_row: list[object] = []
    end_row: list[object] = []
    period_row: list[object] = []
    currency_row: list[object] = []
    header_row: list[object] = []
    value_row_one: list[object] = []
    value_row_two: list[object] = []

    for display_name, ticker in zip(display_names, tickers):
        name_row.extend([None, display_name, None])
        security_row.extend(["Security", ticker, None])
        start_row.extend(["Start Date", 40544, None])
        end_row.extend(["End Date", 46126, None])
        period_row.extend(["Period", "D", None])
        currency_row.extend(["Currency", "USD", None])
        header_row.extend(["Date", "PX_LAST", None])
        value_row_one.extend([46126, 110.0, None])
        value_row_two.extend([46125, 100.0, None])

    rows = [
        [],
        [None, "END DATE", 46126],
        [],
        name_row,
        [],
        security_row,
        start_row,
        end_row,
        period_row,
        currency_row,
        [],
        header_row,
        value_row_one,
        value_row_two,
    ]
    for row in rows:
        sheet.append(row)


def test_parse_assets_workbook_extracts_all_requested_asset_classes() -> None:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        workbook_path = Path(handle.name)

    workbook = Workbook()

    equity = workbook.active
    equity.title = "Equity"
    _append_standard_rows(equity, ["S&P500"], ["SPX Index"])

    commodities = workbook.create_sheet("Commodities")
    _append_standard_rows(commodities, ["Copper"], ["LMCADY LME Comdty"])

    bonds = workbook.create_sheet("Bonds")
    _append_standard_rows(bonds, ["Global Aggregate", "US Lev Loan Index"], ["LEGATRUU Index", "I38932US Index"])

    etfs = workbook.create_sheet("ETFs")
    _append_standard_rows(etfs, ["SPDR S&P 500 ETF Trust", "VanEck Gold Miners ETF"], ["SPY US Equity", "GDX US Equity"])

    crypto = workbook.create_sheet("Crypto")
    _append_standard_rows(crypto, ["Bitcoin"], ["XBT Curncy"])

    top_stocks = workbook.create_sheet("Top-10 stocks")
    _append_standard_rows(top_stocks, ["Amazon"], ["AMZN US Equity"])

    workbook.save(workbook_path)

    try:
        asset_master, prices = parse_assets_workbook(workbook_path)
    finally:
        workbook_path.unlink(missing_ok=True)

    assert set(asset_master["asset_class"]) == {"Equities", "Commodities", "Bonds", "ETFs", "Crypto", "Top-10 Stocks"}
    assert len(asset_master) == 7
    assert len(prices) == 14

    assert "I38932US Index" not in asset_master["bbg_ticker"].tolist()

    spy = asset_master.loc[asset_master["bbg_ticker"] == "SPY US Equity"].iloc[0]
    assert spy["sub_asset_class"] == "Broad Equity ETFs"

    gdx = asset_master.loc[asset_master["bbg_ticker"] == "GDX US Equity"].iloc[0]
    assert gdx["commodity_category"] == "Precious metals"

    bitcoin = asset_master.loc[asset_master["bbg_ticker"] == "XBT Curncy"].iloc[0]
    assert bitcoin["asset_class"] == "Crypto"
    assert bitcoin["series_type"] == "crypto_spot"

    amazon = asset_master.loc[asset_master["bbg_ticker"] == "AMZN US Equity"].iloc[0]
    assert amazon["asset_class"] == "Top-10 Stocks"
    assert amazon["sector_name"] == "Consumer Internet"
